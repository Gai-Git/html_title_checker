import os
from bs4 import BeautifulSoup
from openpyxl import Workbook

def find_empty_and_missing_titles(directory):
    empty_title_files = []
    missing_title_files = []

    for root, _, files in os.walk(directory):
        for file in files:
            if file.endswith(".html") or file.endswith(".htm"):
                file_path = os.path.join(root, file)
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                        
                        # Check if <!doctype html> is present
                        if '<!doctype html>' in content.lower():
                            soup = BeautifulSoup(content, 'html.parser')
                            title = soup.title
                            if title is None:
                                missing_title_files.append(file_path)
                            elif title.string is None or not title.string.strip():
                                empty_title_files.append(file_path)
                except Exception as e:
                    print(f"Error processing file {file_path}: {e}")

    return empty_title_files, missing_title_files

def write_to_excel(empty_title_files, missing_title_files, output_file):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "HTML Title Check"

    # Add headers
    sheet.cell(row=1, column=1, value="Empty Title Files")
    sheet.cell(row=1, column=2, value="Missing Title Files")

    # Write empty title files to the first column starting from row 2
    for idx, file_path in enumerate(empty_title_files, start=2):
        sheet.cell(row=idx, column=1, value=file_path)

    # Write missing title files to the second column starting from row 2
    for idx, file_path in enumerate(missing_title_files, start=2):
        sheet.cell(row=idx, column=2, value=file_path)

    workbook.save(output_file)

if __name__ == "__main__":
    directory = 'C:/Users/gaish/Documents/Intramart-replace/CCFlow/wwwroot'  # Prompt user for directory
    output_file = 'html_title_check_results.xlsx'
    
    empty_title_files, missing_title_files = find_empty_and_missing_titles(directory)
    write_to_excel(empty_title_files, missing_title_files, output_file)
    
    print(f"检测完成，共找到 {len(empty_title_files)} 个标题为空的文件，{len(missing_title_files)} 个缺少标题的文件，已保存到 {output_file}")
